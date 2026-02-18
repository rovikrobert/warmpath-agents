"""Investor report export — generates Excel workbook from agent reports.

Reads latest agent reports from finance_team/reports/ and produces
a 4-sheet Excel workbook for investor meetings.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from io import BytesIO

logger = logging.getLogger(__name__)


def _load_latest_report(name: str) -> dict:
    """Load a cached agent report from finance_team/reports/."""
    from finance_team.shared.config import REPORTS_DIR

    path = REPORTS_DIR / f"{name}_latest.json"
    try:
        return json.loads(path.read_text(errors="replace"))
    except (OSError, json.JSONDecodeError):
        return {}


def generate_workbook() -> bytes:
    """Generate investor report as Excel bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    fm = _load_latest_report("finance_manager")
    cm = _load_latest_report("credits_manager")
    ir = _load_latest_report("investor_relations")
    lc = _load_latest_report("legal_compliance")

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    def _add_header(ws, row, col, text):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    def _add_kv(ws, row, key, value):
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value) if value is not None else "N/A")

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    ws.cell(row=1, column=1, value="WarmPath — Investor Summary").font = Font(
        bold=True, size=14
    )
    ws.cell(
        row=2,
        column=1,
        value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
    )

    row = 4
    _add_header(ws, row, 1, "Metric")
    _add_header(ws, row, 2, "Value")

    ir_metrics = ir.get("metrics", {})
    row += 1
    _add_kv(ws, row, "Test Count", ir_metrics.get("test_count_actual", "N/A"))
    row += 1
    _add_kv(ws, row, "Deployment", "Railway (live)")
    row += 1
    _add_kv(ws, row, "Database Tables", "30")
    row += 1
    _add_kv(ws, row, "Agent Teams", "6 + CoS")

    # --- Sheet 2: Financial Health ---
    ws2 = wb.create_sheet("Financial Health")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 40

    _add_header(ws2, 1, 1, "Metric")
    _add_header(ws2, 1, 2, "Value")

    fm_metrics = fm.get("metrics", {})
    cm_metrics = cm.get("metrics", {})
    row = 2
    for label, key, src in [
        ("Webhook Event Coverage", "webhook_event_coverage", fm_metrics),
        ("Billing Completeness", "billing_completeness_score", fm_metrics),
        ("Credit Economy Integrity", "credit_economy_integrity", cm_metrics),
        ("Monthly Burn Rate", "monthly_burn_rate", fm_metrics),
        ("Cash Runway (months)", "cash_runway_months", fm_metrics),
        ("Credit Velocity (days)", "credit_velocity_days", cm_metrics),
        ("Credit Gini", "credit_gini", cm_metrics),
        ("Expiry Rate", "credit_expiry_rate", cm_metrics),
    ]:
        val = src.get(key, "N/A")
        if isinstance(val, float) and key in (
            "webhook_event_coverage",
            "billing_completeness_score",
            "credit_economy_integrity",
        ):
            val = f"{val:.0%}"
        elif isinstance(val, float) and key == "monthly_burn_rate":
            val = f"${val}"
        _add_kv(ws2, row, label, val)
        row += 1

    # --- Sheet 3: Compliance ---
    ws3 = wb.create_sheet("Compliance")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 40

    _add_header(ws3, 1, 1, "Metric")
    _add_header(ws3, 1, 2, "Value")

    lc_metrics = lc.get("metrics", {})
    row = 2
    _add_kv(ws3, row, "Compliance Score", lc_metrics.get("compliance_score", "N/A"))
    row += 1
    _add_kv(ws3, row, "Critical Findings", lc_metrics.get("critical_findings", "N/A"))
    row += 1
    sec_score = lc_metrics.get("security_compliance_score", 0)
    _add_kv(
        ws3,
        row,
        "Security Score",
        f"{sec_score:.0%}" if isinstance(sec_score, float) else sec_score,
    )
    row += 1
    _add_kv(
        ws3,
        row,
        "Consent Gates",
        f"{lc_metrics.get('consent_gates_found', 0)}/{lc_metrics.get('consent_gates_expected', 0)}",
    )
    row += 1
    _add_kv(
        ws3,
        row,
        "GDPR Deletion Functions",
        f"{lc_metrics.get('gdpr_deletion_functions_found', 0)}/{lc_metrics.get('gdpr_deletion_functions_expected', 0)}",
    )
    row += 1
    _add_kv(
        ws3,
        row,
        "Suppression List",
        "Active" if lc_metrics.get("has_suppression_model") else "Missing",
    )
    row += 1
    _add_kv(
        ws3,
        row,
        "Deletion Verification",
        "Active"
        if lc_metrics.get("deletion_verification_available")
        else "DB required",
    )

    # --- Sheet 4: Technical Readiness ---
    ws4 = wb.create_sheet("Technical Readiness")
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 40

    _add_header(ws4, 1, 1, "Metric")
    _add_header(ws4, 1, 2, "Value")

    row = 2
    _add_kv(ws4, row, "Test Count", ir_metrics.get("test_count_actual", "N/A"))
    row += 1
    _add_kv(ws4, row, "TODO/FIXME Count", ir_metrics.get("todo_fixme_count", "N/A"))
    row += 1
    _add_kv(
        ws4,
        row,
        "Schema Maturity",
        ir_metrics.get("schema_maturity_score", "N/A"),
    )
    row += 1
    _add_kv(
        ws4,
        row,
        "Token Version (JWT)",
        "Yes" if lc_metrics.get("has_token_version") else "No",
    )
    row += 1
    _add_kv(
        ws4,
        row,
        "Account Lockout",
        "Yes" if lc_metrics.get("has_locked_until") else "No",
    )
    row += 1
    _add_kv(
        ws4,
        row,
        "Security Headers",
        "Yes" if lc_metrics.get("has_security_headers_middleware") else "No",
    )
    row += 1
    _add_kv(
        ws4,
        row,
        "Audit Logs",
        "Yes" if lc_metrics.get("has_audit_logs") else "No",
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
